"""Generate a formatted Excel trade report from the OBI trade log CSV."""
import csv
import os
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side as XlSide
from openpyxl.utils import get_column_letter

LOG_DIR = "data/logs"
OUT = "data/logs/trade_report.xlsx"

# Find the latest trades CSV
csvs = sorted(glob.glob(os.path.join(LOG_DIR, "trades_*.csv")))
if not csvs:
    print("No trade logs found in data/logs/")
    exit(1)

csv_path = csvs[-1]
print(f"Reading: {csv_path}")

rows = []
with open(csv_path, 'r') as f:
    reader = csv.reader(f)
    for r in reader:
        rows.append(r)

if len(rows) < 2:
    print("No trades logged yet.")
    exit(1)

header = rows[0]
data = rows[1:]

wb = Workbook()

# Colors
HDR_FILL = PatternFill('solid', fgColor='1B2A4A')
HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
WIN_FILL = PatternFill('solid', fgColor='E8F5E9')
LOSS_FILL = PatternFill('solid', fgColor='FFEBEE')
ENTRY_FILL = PatternFill('solid', fgColor='E3F2FD')
EXIT_FILL = PatternFill('solid', fgColor='FFF3E0')
SUMM_FILL = PatternFill('solid', fgColor='F3E5F5')
DATA_FONT = Font(name='Arial', size=10)
MONEY_FONT_GREEN = Font(name='Arial', size=10, color='2E7D32', bold=True)
MONEY_FONT_RED = Font(name='Arial', size=10, color='C62828', bold=True)
TITLE_FONT = Font(name='Arial', size=14, bold=True, color='1B2A4A')
STAT_LABEL = Font(name='Arial', size=11, color='666666')
STAT_VALUE = Font(name='Arial', size=11, bold=True, color='1B2A4A')
thin_border = Border(
    left=XlSide(style='thin', color='DDDDDD'),
    right=XlSide(style='thin', color='DDDDDD'),
    top=XlSide(style='thin', color='DDDDDD'),
    bottom=XlSide(style='thin', color='DDDDDD'),
)

# ── SHEET 1: SUMMARY ──
ws_sum = wb.active
ws_sum.title = "Summary"
ws_sum.sheet_properties.tabColor = "1B2A4A"

# Calculate stats
entries = [r for r in data if len(r) > 12 and r[12] == 'ENTRY']
exits = [r for r in data if len(r) > 12 and r[12] == 'EXIT']
summaries = [r for r in data if len(r) > 12 and r[12] == 'SUMMARY']

pnl_col = 10  # pnl column index
pnls = []
for r in exits:
    try:
        pnls.append(float(r[pnl_col]))
    except (ValueError, IndexError):
        pass

total_trades = len(exits)
wins = len([p for p in pnls if p > 0])
losses = len([p for p in pnls if p <= 0])
total_pnl = sum(pnls)
avg_pnl = total_pnl / max(total_trades, 1)
best_trade = max(pnls) if pnls else 0
worst_trade = min(pnls) if pnls else 0
win_rate = wins / max(total_trades, 1)

ws_sum['A1'] = 'OBI Paper Trading Report'
ws_sum['A1'].font = TITLE_FONT
ws_sum.merge_cells('A1:D1')

ws_sum['A2'] = f'Generated from: {os.path.basename(csv_path)}'
ws_sum['A2'].font = Font(name='Arial', size=10, color='999999')

stats = [
    ('Total Trades', total_trades),
    ('Wins', wins),
    ('Losses', losses),
    ('Win Rate', f'{win_rate:.1%}'),
    ('Total PnL', f'${total_pnl:+.2f}'),
    ('Avg PnL/Trade', f'${avg_pnl:+.2f}'),
    ('Best Trade', f'${best_trade:+.2f}'),
    ('Worst Trade', f'${worst_trade:+.2f}'),
    ('Signals → Entries', f'{len(entries)}'),
]

for i, (label, value) in enumerate(stats):
    row = i + 4
    ws_sum[f'A{row}'] = label
    ws_sum[f'A{row}'].font = STAT_LABEL
    ws_sum[f'B{row}'] = value
    ws_sum[f'B{row}'].font = STAT_VALUE
    if 'PnL' in label or 'Trade' in label:
        if isinstance(value, str) and '+' in value:
            ws_sum[f'B{row}'].font = MONEY_FONT_GREEN
        elif isinstance(value, str) and '-' in value:
            ws_sum[f'B{row}'].font = MONEY_FONT_RED

ws_sum.column_dimensions['A'].width = 18
ws_sum.column_dimensions['B'].width = 15

# ── SHEET 2: ALL TRADES ──
ws_trades = wb.create_sheet("Trades")
ws_trades.sheet_properties.tabColor = "2196F3"

col_widths = [20, 45, 8, 8, 8, 10, 10, 10, 45, 15, 10, 8, 8]
for i, w in enumerate(col_widths):
    ws_trades.column_dimensions[get_column_letter(i + 1)].width = w

# Header row
for j, h in enumerate(header):
    cell = ws_trades.cell(row=1, column=j + 1, value=h.upper())
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Data rows
for i, row_data in enumerate(data):
    row_num = i + 2
    row_type = row_data[12] if len(row_data) > 12 else ''

    for j, val in enumerate(row_data):
        cell = ws_trades.cell(row=row_num, column=j + 1, value=val)
        cell.font = DATA_FONT
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

        # Color code by type
        if row_type == 'ENTRY':
            cell.fill = ENTRY_FILL
        elif row_type == 'EXIT':
            # Color by PnL
            try:
                pnl_val = float(row_data[pnl_col])
                cell.fill = WIN_FILL if pnl_val > 0 else LOSS_FILL
                if j == pnl_col:
                    cell.font = MONEY_FONT_GREEN if pnl_val > 0 else MONEY_FONT_RED
            except (ValueError, IndexError):
                cell.fill = EXIT_FILL
        elif row_type == 'SUMMARY':
            cell.fill = SUMM_FILL

# Freeze top row
ws_trades.freeze_panes = 'A2'

# ── SHEET 3: PNL TRACKER ──
ws_pnl = wb.create_sheet("PnL Tracker")
ws_pnl.sheet_properties.tabColor = "4CAF50"

pnl_headers = ['#', 'Time', 'Market', 'Token', 'Side', 'Entry', 'Exit', 'PnL', 'Cumulative PnL', 'Result']
for j, h in enumerate(pnl_headers):
    cell = ws_pnl.cell(row=1, column=j + 1, value=h)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

cumulative = 0
trade_num = 0

# Match entries with exits by pairing them
entry_stack = []
for row_data in data:
    row_type = row_data[12] if len(row_data) > 12 else ''
    if row_type == 'ENTRY':
        entry_stack.append(row_data)
    elif row_type == 'EXIT' and entry_stack:
        entry = entry_stack.pop(0)
        trade_num += 1
        try:
            pnl_val = float(row_data[pnl_col])
        except (ValueError, IndexError):
            pnl_val = 0
        cumulative += pnl_val

        r = trade_num + 1
        ws_pnl.cell(row=r, column=1, value=trade_num).font = DATA_FONT
        ws_pnl.cell(row=r, column=2, value=row_data[0]).font = DATA_FONT
        ws_pnl.cell(row=r, column=3, value=entry[1][:30]).font = DATA_FONT
        ws_pnl.cell(row=r, column=4, value=entry[2]).font = DATA_FONT
        ws_pnl.cell(row=r, column=5, value=entry[3]).font = DATA_FONT
        ws_pnl.cell(row=r, column=6, value=entry[5]).font = DATA_FONT
        ws_pnl.cell(row=r, column=7, value=row_data[5]).font = DATA_FONT

        pnl_cell = ws_pnl.cell(row=r, column=8, value=f'${pnl_val:+.2f}')
        pnl_cell.font = MONEY_FONT_GREEN if pnl_val > 0 else MONEY_FONT_RED

        cum_cell = ws_pnl.cell(row=r, column=9, value=f'${cumulative:+.2f}')
        cum_cell.font = MONEY_FONT_GREEN if cumulative >= 0 else MONEY_FONT_RED
        cum_cell.fill = WIN_FILL if cumulative >= 0 else LOSS_FILL

        result_cell = ws_pnl.cell(row=r, column=10, value='WIN ✓' if pnl_val > 0 else 'LOSS ✗')
        result_cell.font = MONEY_FONT_GREEN if pnl_val > 0 else MONEY_FONT_RED
        result_cell.alignment = Alignment(horizontal='center')

        for j in range(1, 11):
            ws_pnl.cell(row=r, column=j).border = thin_border

pnl_widths = [5, 20, 32, 8, 8, 10, 10, 12, 15, 10]
for i, w in enumerate(pnl_widths):
    ws_pnl.column_dimensions[get_column_letter(i + 1)].width = w

ws_pnl.freeze_panes = 'A2'

wb.save(OUT)
print(f"\n✓ Report saved to: {OUT}")
print(f"  Trades: {total_trades} | Wins: {wins} | Losses: {losses} | PnL: ${total_pnl:+.2f}")
