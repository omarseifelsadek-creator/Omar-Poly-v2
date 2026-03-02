"""
generate_pair_report.py — PolyQuant Professional Excel Report

Generates hourly reports every 12 windows (1 hour of 5-min trading).

File naming: PolyQuant_BTC5m_{Date}_{StartTime}_to_{EndTime}.xlsx
Two tabs:
  1. Window_Summaries — one row per 5-min window (macro view)
  2. Execution_Log    — every individual fill (micro view)

Can also be called standalone to generate from CSV data.
"""

import os
import csv
import glob
import time
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side as XlSide
from openpyxl.utils import get_column_letter


LOG_DIR = "data/logs"

# ══════════════════════════════════════════════════════
# STYLE CONSTANTS
# ══════════════════════════════════════════════════════

NAVY = '0D1B2A'
DARK_BLUE = '1B2A4A'
WHITE = 'FFFFFF'
GREEN_BG = 'E8F5E9'
RED_BG = 'FFEBEE'
GREY_BG = 'F5F5F5'
AMBER_BG = 'FFF8E1'
YES_BG = 'E8F5E9'
NO_BG = 'E3F2FD'
PANIC_BG = 'FFCDD2'

TITLE_FONT = Font(name='Arial', size=14, bold=True, color=NAVY)
SUBTITLE_FONT = Font(name='Arial', size=10, color='888888', italic=True)
HDR_FONT = Font(name='Arial', size=10, bold=True, color=WHITE)
DATA_FONT = Font(name='Arial', size=10, color='333333')
BOLD_DATA = Font(name='Arial', size=10, bold=True, color='333333')
LABEL_FONT = Font(name='Arial', size=10, color='666666')
MONEY_GREEN = Font(name='Arial', size=10, color='2E7D32', bold=True)
MONEY_RED = Font(name='Arial', size=10, color='C62828', bold=True)
KPI_VALUE = Font(name='Arial', size=13, bold=True, color=NAVY)
KPI_LABEL = Font(name='Arial', size=9, color='999999')
KPI_GREEN = Font(name='Arial', size=13, bold=True, color='2E7D32')
KPI_RED = Font(name='Arial', size=13, bold=True, color='C62828')

HDR_FILL = PatternFill('solid', fgColor=DARK_BLUE)
WIN_FILL = PatternFill('solid', fgColor=GREEN_BG)
LOSS_FILL = PatternFill('solid', fgColor=RED_BG)
NEUTRAL_FILL = PatternFill('solid', fgColor=AMBER_BG)
KPI_FILL = PatternFill('solid', fgColor='F0F4F8')
TOTALS_FILL = PatternFill('solid', fgColor='E0E0E0')

thin_border = Border(
    left=XlSide(style='thin', color='DDDDDD'),
    right=XlSide(style='thin', color='DDDDDD'),
    top=XlSide(style='thin', color='DDDDDD'),
    bottom=XlSide(style='thin', color='DDDDDD'),
)
thick_bottom = Border(bottom=XlSide(style='thick', color=NAVY))
CENTER = Alignment(horizontal='center', vertical='center')


# ══════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════

def _sf(val, default=0.0):
    try:
        return float(val)
    except (ValueError, TypeError):
        return default

def _si(val, default=0):
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default

def _money_font(val):
    return MONEY_GREEN if val >= 0 else MONEY_RED

def _kpi_font(val):
    return KPI_GREEN if val >= 0 else KPI_RED

def _header_row(ws, row, headers, widths):
    for j, h in enumerate(headers):
        cell = ws.cell(row=row, column=j + 1, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = thin_border
    for j, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(j + 1)].width = w

def _data_cell(ws, row, col, val, font=None, fill=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = font or DATA_FONT
    cell.border = thin_border
    cell.alignment = CENTER
    if fill:
        cell.fill = fill
    return cell


# ══════════════════════════════════════════════════════
# GENERATE FROM IN-MEMORY DATA (called by _auto_report)
# ══════════════════════════════════════════════════════

def generate_from_memory(window_records: list, fill_records: list,
                         session_pnl: float, output_path: str) -> str:
    """
    Generate report from in-memory data arrays.

    Args:
        window_records: List of dicts, one per settled window
        fill_records: List of dicts, one per individual fill
        session_pnl: Cumulative session PnL for header
        output_path: Where to save the .xlsx

    Returns:
        Path to saved file.
    """
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb = Workbook()
    _build_window_summaries(wb, window_records, fill_records, session_pnl)
    _build_execution_log(wb, fill_records)
    wb.save(output_path)
    return output_path


# ══════════════════════════════════════════════════════
# TAB 1: WINDOW SUMMARIES
# ══════════════════════════════════════════════════════

def _build_window_summaries(wb, windows, fills, session_pnl):
    ws = wb.active
    ws.title = "Window_Summaries"
    ws.sheet_properties.tabColor = DARK_BLUE

    # Title
    ws.merge_cells('A1:M1')
    ws['A1'] = '⚡ PolyQuant BTC 5m — Window Performance'
    ws['A1'].font = TITLE_FONT

    # Subtitle
    n_win = len(windows)
    wins = len([w for w in windows if _sf(w.get('net_pnl')) > 0])
    total_pnl = sum(_sf(w.get('net_pnl')) for w in windows)
    total_pairs = sum(_si(w.get('completed_pairs')) for w in windows)

    ws.merge_cells('A2:M2')
    ws['A2'] = (
        f'{time.strftime("%B %d, %Y %H:%M")}  •  '
        f'{n_win} windows  •  Session PnL: ${session_pnl:+.2f}'
    )
    ws['A2'].font = SUBTITLE_FONT

    # KPI row
    kpi_row = 4
    for c_lbl, c_val, label, value, raw in [
        ('A', 'B', 'Report PnL', f'${total_pnl:+.2f}', total_pnl),
        ('D', 'E', 'Win Rate', f'{wins}/{n_win} ({wins/max(n_win,1):.0%})', 1),
        ('G', 'H', 'Pairs', f'{total_pairs:,}', 1),
    ]:
        ws[f'{c_lbl}{kpi_row}'] = label
        ws[f'{c_lbl}{kpi_row}'].font = KPI_LABEL
        ws[f'{c_val}{kpi_row}'] = value
        ws[f'{c_val}{kpi_row}'].font = _kpi_font(raw) if '$' in value else KPI_VALUE
        for c in [c_lbl, c_val]:
            ws[f'{c}{kpi_row}'].fill = KPI_FILL
            ws[f'{c}{kpi_row}'].border = thin_border

    # Data table
    headers = [
        'Window Start', 'Window End', 'Winner', 'Pairs',
        'Avg Pair Cost', 'Panic Hedges', 'Capital',
        'Net PnL', 'Rejection Rate', 'Dead Zone Blocks',
        'Cap Exhausted At', 'Max Unhedged ($)', 'Avg Hedge Time (s)',
    ]
    widths = [18, 18, 9, 8, 14, 14, 12, 12, 14, 17, 16, 16, 18]

    data_start = 6
    _header_row(ws, data_start, headers, widths)

    for i, w in enumerate(windows):
        r = data_start + 1 + i
        pnl = _sf(w.get('net_pnl'))
        row_fill = WIN_FILL if pnl > 0 else LOSS_FILL if pnl < 0 else NEUTRAL_FILL

        attempted = _si(w.get('fills_attempted'))
        rejected = _si(w.get('fills_rejected'))
        rej_rate = f'{rejected/attempted:.0%}' if attempted > 0 else '0%'

        wid = w.get('window_id', '')
        panic_count = len([
            f for f in fills
            if f.get('window_id') == wid and f.get('zone') == 'Panic'
        ])

        vals = [
            (w.get('window_start_time', ''), DATA_FONT),
            (w.get('window_end_time', ''), DATA_FONT),
            (w.get('winner', ''), BOLD_DATA),
            (_si(w.get('completed_pairs')), BOLD_DATA),
            (f"${_sf(w.get('avg_pair_cost')):.4f}", DATA_FONT),
            (panic_count, MONEY_RED if panic_count > 0 else DATA_FONT),
            (f"${_sf(w.get('total_capital')):.2f}", DATA_FONT),
            (f"${pnl:+.2f}", _money_font(pnl)),
            (rej_rate, DATA_FONT),
            (_si(w.get('dead_zone_blocks')), DATA_FONT),
            (w.get('capital_exhaustion_time', 'N/A'), DATA_FONT),
            (f"${_sf(w.get('max_unhedged_exposure', '0')):.2f}", DATA_FONT),
            (w.get('avg_time_to_hedge', 'N/A'), DATA_FONT),
        ]

        for j, (val, font) in enumerate(vals):
            _data_cell(ws, r, j + 1, val, font=font, fill=row_fill)

    # Totals row
    if windows:
        tot_r = data_start + 1 + len(windows)
        _data_cell(ws, tot_r, 1, 'TOTALS', font=BOLD_DATA, fill=TOTALS_FILL)
        for j in range(2, 14):
            _data_cell(ws, tot_r, j, '', fill=TOTALS_FILL)
        _data_cell(ws, tot_r, 4, total_pairs, font=BOLD_DATA, fill=TOTALS_FILL)
        total_cap = sum(_sf(w.get('total_capital')) for w in windows)
        _data_cell(ws, tot_r, 7, f'${total_cap:.2f}', font=BOLD_DATA, fill=TOTALS_FILL)
        _data_cell(ws, tot_r, 8, f'${total_pnl:+.2f}', font=_money_font(total_pnl), fill=TOTALS_FILL)

    ws.freeze_panes = f'A{data_start + 1}'


# ══════════════════════════════════════════════════════
# TAB 2: EXECUTION LOG
# ══════════════════════════════════════════════════════

def _build_execution_log(wb, fills):
    ws = wb.create_sheet("Execution_Log")
    ws.sheet_properties.tabColor = '4CAF50'

    ws.merge_cells('A1:M1')
    ws['A1'] = '📋 Execution Log — Every Fill'
    ws['A1'].font = TITLE_FONT

    total_slip = 0
    sniper_n = value_n = panic_n = 0
    for f in fills:
        q = _sf(f.get('quoted_ask'))
        v = _sf(f.get('vwap_fill'))
        total_slip += (v - q) * 100
        z = f.get('zone', '')
        if z == 'Sniper': sniper_n += 1
        elif z == 'Value': value_n += 1
        elif z == 'Panic': panic_n += 1

    avg_slip = total_slip / len(fills) if fills else 0

    ws.merge_cells('A2:M2')
    ws['A2'] = (
        f'{len(fills)} fills  •  '
        f'Sniper: {sniper_n}  Value: {value_n}  Panic: {panic_n}  •  '
        f'Avg Slippage: {avg_slip:+.2f}¢'
    )
    ws['A2'].font = SUBTITLE_FONT

    headers = [
        'Timestamp', 'Window', 'Token', 'Shares',
        'Zone', 'Quoted Ask', 'VWAP Fill', 'Slippage (¢)', 'Fee %',
        'Opp Leg Ask', 'Ask Age (ms)', 'OBI', 'Time-to-Hedge (s)',
    ]
    widths = [14, 14, 7, 8, 9, 12, 12, 12, 8, 12, 13, 8, 18]

    data_start = 4
    _header_row(ws, data_start, headers, widths)

    for i, f in enumerate(fills):
        r = data_start + 1 + i
        token = f.get('token', '?')
        zone = f.get('zone', '?')
        quoted = _sf(f.get('quoted_ask'))
        vwap = _sf(f.get('vwap_fill'))
        slippage = (vwap - quoted) * 100

        if zone == 'Panic':
            row_fill = PatternFill('solid', fgColor=PANIC_BG)
        elif i % 2 == 0:
            row_fill = PatternFill('solid', fgColor=YES_BG if token == 'YES' else NO_BG)
        else:
            row_fill = PatternFill('solid', fgColor=GREY_BG)

        slip_font = MONEY_RED if slippage > 0.5 else MONEY_GREEN if slippage <= 0 else DATA_FONT

        vals = [
            (f.get('timestamp', ''), DATA_FONT),
            (f.get('window_id', ''), DATA_FONT),
            (token, BOLD_DATA),
            (_sf(f.get('shares')), DATA_FONT),
            (zone, BOLD_DATA),
            (f'${quoted:.4f}', DATA_FONT),
            (f'${vwap:.4f}', DATA_FONT),
            (f'{slippage:+.2f}', slip_font),
            (f.get('fee_pct', ''), DATA_FONT),
            (f.get('opposite_leg_ask', 'N/A'), DATA_FONT),
            (f.get('ask_age_ms', 'N/A'), DATA_FONT),
            (f.get('obi_ratio', 'N/A'), DATA_FONT),
            (f.get('time_to_hedge_sec', 'N/A'), DATA_FONT),
        ]

        for j, (val, font) in enumerate(vals):
            _data_cell(ws, r, j + 1, val, font=font, fill=row_fill)

    ws.freeze_panes = f'A{data_start + 1}'


# ══════════════════════════════════════════════════════
# STANDALONE: GENERATE FROM CSV
# ══════════════════════════════════════════════════════

def generate_pair_report(output_path: str = None) -> str:
    """Generate report from CSV files on disk."""
    window_files = sorted(glob.glob(os.path.join(LOG_DIR, "pair_windows_*.csv")))
    windows = []
    for fp in window_files:
        try:
            with open(fp, 'r') as fh:
                for row in csv.DictReader(fh):
                    ts = row.get('timestamp', '')
                    market = row.get('market', '')
                    win_start = ts
                    win_end = ''
                    if '—' in market:
                        time_part = market.split('—')[-1].strip()
                        parts = time_part.replace(' UTC', '').split('-')
                        if len(parts) == 2:
                            win_start = parts[0].strip()
                            win_end = parts[1].strip()

                    windows.append({
                        'window_start_time': win_start,
                        'window_end_time': win_end,
                        'window_id': f'{ts[:10]}_{win_start}',
                        'winner': row.get('winner', ''),
                        'completed_pairs': row.get('completed_pairs', '0'),
                        'avg_pair_cost': row.get('avg_pair_cost', '0'),
                        'total_capital': row.get('total_capital', '0'),
                        'net_pnl': row.get('net_pnl', '0'),
                        'fills_attempted': row.get('num_buys', '0'),
                        'fills_rejected': '0',
                        'dead_zone_blocks': '0',
                    })
        except Exception:
            pass

    buy_files = sorted(glob.glob(os.path.join(LOG_DIR, "pair_buys_*.csv")))
    fills = []
    for fp in buy_files:
        try:
            with open(fp, 'r') as fh:
                for row in csv.DictReader(fh):
                    ask = _sf(row.get('ask_price'))
                    t_rem = _sf(row.get('time_remaining'))
                    if t_rem < 60:
                        zone = 'Panic'
                    elif ask <= 0.35:
                        zone = 'Sniper'
                    elif ask <= 0.44:
                        zone = 'Value'
                    else:
                        zone = 'Panic'

                    fills.append({
                        'timestamp': row.get('timestamp', '')[11:],  # HH:MM:SS
                        'window_id': '',
                        'token': row.get('side', ''),
                        'shares': row.get('qty', '0'),
                        'zone': zone,
                        'quoted_ask': row.get('ask_price', '0'),
                        'vwap_fill': row.get('vwap_price', row.get('ask_price', '0')),
                        'fee_pct': row.get('fee_pct', ''),
                    })
        except Exception:
            pass

    if not windows and not fills:
        print("No pair trading data found in data/logs/")
        return ""

    cum = sum(_sf(w.get('net_pnl')) for w in windows)

    if not output_path:
        output_path = os.path.join(LOG_DIR, "pair_report.xlsx")

    return generate_from_memory(windows, fills, cum, output_path)


if __name__ == "__main__":
    path = generate_pair_report()
    if path:
        print(f"\n✓ Report saved to: {path}")
        window_files = sorted(glob.glob(os.path.join(LOG_DIR, "pair_windows_*.csv")))
        pnls = []
        for f in window_files:
            try:
                with open(f, 'r') as fh:
                    for row in csv.DictReader(fh):
                        pnls.append(_sf(row.get('net_pnl')))
            except Exception:
                pass
        if pnls:
            wins = len([p for p in pnls if p > 0])
            total = sum(pnls)
            print(f"  Windows: {len(pnls)} | Won: {wins} | Lost: {len(pnls)-wins} | PnL: ${total:+.2f}")
